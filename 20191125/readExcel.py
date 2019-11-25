from openpyxl.reader.excel import load_workbook


def readXlsxFile(path):
    file = load_workbook(filename=path)
    print(file.get_sheet_names)
    sheets = file.get_sheet_names()
    sheet = file.get_sheet_by_name(sheets[0])
    for lineNum in range(1, sheet.max_row + 1):
        lineList = []
        #print(sheet.max_row, sheet.max_column)
        for columnNum in range(1, sheet.max_column + 1):
            # 拿数据
            value = sheet.cell(row=lineNum,
                               column=columnNum).value
            if value != None:
                lineList.append(value)
        print(lineList)


path = r"D:\\workspace\\python\\demo\\20191125\\new.xlsx"
readXlsxFile(path)
